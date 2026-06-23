"""
评论保存工具
============
将评论数据保存为 CSV / JSON / Excel 格式。

用法:
  python core/save_json.py BV1k8LX64EpC           ← 默认存 JSON
  python core/save_json.py BV1k8LX64EpC 10 --csv  ← 存 CSV
  python core/save_json.py BV1k8LX64EpC --all     ← 三种格式都存
"""

import sys, os, json, csv, time
from typing import Optional

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from core import bilibili_comment as bili

# 数据目录（插件根目录的 data/）
_PLUGIN_DIR = os.path.dirname(os.path.dirname(__file__))
_DATA_DIR = os.path.join(_PLUGIN_DIR, "data")


def save_comments(bv: str, comments: list,
                  save_csv: bool = True,
                  save_json: bool = True,
                  save_excel: bool = False,
                  data_dir: str = None) -> dict:
    """
    保存评论到文件（统一入口）。

    参数:
        bv:          BV号（用于文件名）
        comments:    评论列表（已爬好的结构化数据）
        save_csv:    是否保存 CSV
        save_json:   是否保存 JSON
        save_excel:  是否保存 Excel（需 openpyxl）
        data_dir:    保存目录（默认 data/）

    返回:
        {"csv": 路径或None, "json": 路径或None, "excel": 路径或None}
    """
    base_dir = data_dir or _DATA_DIR
    ts = int(time.time())
    prefix = f"bili_{bv[:10]}_{ts}"
    result = {}

    if save_csv:
        result["csv"] = _save_csv(comments, base_dir, prefix)
    if save_json:
        result["json"] = _save_json(comments, bv, base_dir, prefix)
    if save_excel:
        result["excel"] = _save_excel(comments, base_dir, prefix)

    return result


# ─── CSV ─────────────────────────────────────────────────────────────

def _save_csv(comments: list, base_dir: str, prefix: str) -> str:
    """保存为 CSV 文件"""
    import stat
    csv_dir = base_dir
    os.makedirs(csv_dir, exist_ok=True)
    path = os.path.join(csv_dir, f"{prefix}.csv")

    fd = os.open(path, os.O_WRONLY | os.O_CREAT | os.O_TRUNC,
                 stat.S_IRUSR | stat.S_IWUSR | stat.S_IRGRP | stat.S_IROTH)
    with os.fdopen(fd, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["层级", "rpid", "用户", "UID", "评论内容", "点赞数", "时间", "所属主评论"])
        for c in comments:
            writer.writerow([
                "主评论",
                c.get("rpid", ""),
                c.get("user", ""),
                c.get("uid", ""),
                c.get("content", "").replace("\n", " "),
                c.get("likes", 0),
                c.get("time_str", ""),
                "",
            ])
            for sub in c.get("sub_replies", []):
                writer.writerow([
                    "子回复",
                    sub.get("rpid", ""),
                    sub.get("user", ""),
                    sub.get("uid", ""),
                    sub.get("content", "").replace("\n", " "),
                    sub.get("likes", 0),
                    sub.get("time_str", ""),
                    c.get("rpid", ""),
                ])
    return path


# ─── JSON ────────────────────────────────────────────────────────────

def _save_json(comments: list, bv: str, base_dir: str, prefix: str) -> str:
    """保存为 JSON 文件"""
    import stat
    os.makedirs(base_dir, exist_ok=True)
    path = os.path.join(base_dir, f"{prefix}.json")

    output = {
        "video": bv,
        "crawled_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "main_count": len(comments),
        "comments": comments,
    }

    fd = os.open(path, os.O_WRONLY | os.O_CREAT | os.O_TRUNC,
                 stat.S_IRUSR | stat.S_IWUSR | stat.S_IRGRP | stat.S_IROTH)
    with os.fdopen(fd, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    return path


# ─── Excel ───────────────────────────────────────────────────────────

def _save_excel(comments: list, base_dir: str, prefix: str) -> Optional[str]:
    """保存为 Excel 文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[警告] 未安装 openpyxl，无法保存 Excel")
        return None

    os.makedirs(base_dir, exist_ok=True)
    path = os.path.join(base_dir, f"{prefix}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "评论"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    headers = ["层级", "rpid", "用户", "UID", "评论内容", "点赞数", "时间", "所属主评论"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for c in comments:
        ws.cell(row, 1, "主评论")
        ws.cell(row, 2, c.get("rpid", ""))
        ws.cell(row, 3, c.get("user", ""))
        ws.cell(row, 4, c.get("uid", ""))
        ws.cell(row, 5, c.get("content", ""))
        ws.cell(row, 6, c.get("likes", 0))
        ws.cell(row, 7, c.get("time_str", ""))
        ws.cell(row, 8, "")
        row += 1
        for sub in c.get("sub_replies", []):
            ws.cell(row, 1, "子回复")
            ws.cell(row, 2, sub.get("rpid", ""))
            ws.cell(row, 3, sub.get("user", ""))
            ws.cell(row, 4, sub.get("uid", ""))
            ws.cell(row, 5, sub.get("content", ""))
            ws.cell(row, 6, sub.get("likes", 0))
            ws.cell(row, 7, sub.get("time_str", ""))
            ws.cell(row, 8, c.get("rpid", ""))
            row += 1

    # 自适应列宽
    for col in range(1, 9):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    wb.save(path)
    return path


# ─── 旧接口保留（兼容） ────────────────────────────────────────────

def save_comments_to_json(bv: str, max_pages: int = 200) -> str:
    """旧接口：爬取 + 存 JSON（保留兼容）"""
    ok, data, main_count, sub_count = bili.get_comments_all(bv, max_main_pages=max_pages)
    if not ok:
        raise RuntimeError(f"爬取失败: {data}")
    ok, json_str, mc, sc = bili.get_comments_all(bv, max_main_pages=max_pages, output_format="json")
    if not ok:
        raise RuntimeError(f"JSON 生成失败: {json_str}")
    comments = json.loads(json_str)
    result = save_comments(bv, comments, save_csv=False, save_json=True, save_excel=False)
    return result["json"]


# ─── 独立运行 ────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="B站评论保存工具")
    parser.add_argument("bv", help="BV号")
    parser.add_argument("pages", nargs="?", type=int, default=200, help="页数")
    parser.add_argument("--csv", action="store_true", help="保存 CSV")
    parser.add_argument("--json", action="store_true", help="保存 JSON")
    parser.add_argument("--excel", action="store_true", help="保存 Excel")
    parser.add_argument("--all", action="store_true", help="三种格式都保存")
    args = parser.parse_args()

    if args.all:
        args.csv = args.json = args.excel = True
    if not any([args.csv, args.json, args.excel]):
        args.json = True

    print(f"正在爬取 {args.bv} ({args.pages}页)...")
    ok, data, mc, sc = bili.get_comments_all(args.bv, max_main_pages=args.pages)
    if not ok:
        print(f"爬取失败: {data}")
        sys.exit(1)

    ok, json_str, mc, sc = bili.get_comments_all(args.bv, max_main_pages=args.pages, output_format="json")
    if not ok:
        print(f"JSON 获取失败: {json_str}")
        sys.exit(1)

    comments = json.loads(json_str)
    result = save_comments(args.bv, comments,
                           save_csv=args.csv, save_json=args.json, save_excel=args.excel)

    for fmt, path in result.items():
        if path:
            print(f"[{fmt.upper()}] {path}")
